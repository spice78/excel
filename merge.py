#!python
#from ctypes import alignment
#from msilib.schema import Font
from lib2to3.pgen2.token import SLASH
import openpyxl as xl
from tkinter import *
from tkinter.filedialog import askopenfilename

COPY_COLS = 3  
#FOLDER = 'H:/temp/'
ADD_SLASH = '/'

coun_raz = 1 

base = Tk()
base.title("Мержит файлы в один, по очереди")
#base.resizable(False, False)
base.geometry("600x500")

folder_name = StringVar()

def browsefunc():
    global coun_raz
    folder = folder_name.get()

    filename1 = askopenfilename(filetypes=(("excel files","*.xlsx"),
                                            ("All files","*.*")))
    wb1 = xl.load_workbook(filename1) 
    ws1 = wb1.worksheets[0]
    rows_wb1 = ws1.max_row
    
    filename2 = askopenfilename(filetypes=(("excel files","*.xlsx"),
                                            ("All files","*.*")))
    wb2 = xl.load_workbook(filename2) 
    ws2 = wb2.worksheets[0]
    rows_wb2 = ws2.max_row
    
    i = 1    
    for c in range(1, COPY_COLS+1):
        for i in range(0, rows_wb1):
            ws2.cell(rows_wb2+i+1, c).value = ws1.cell(i+3, c).value
        
    wb1.close    
    wb2.save(folder + ADD_SLASH + "merged.xlsx")
    wb2.close

    folder_name.set(folder)

    coun_raz += 1
    
    lab_coun['text'] = coun_raz

but_first = Button(
            base, 
            text="Сначала файл n+1 потом n = 1 \n после (n+i) и merged. i += 1",
            bg='black', fg='white', width=30, height=5, font=('Times New Roman', 18), 
            justify='left', command=browsefunc
            )
lab_coun = Label(
            base,
            width=15, height=5, bg='white', fg='black',
            font=('Times New Roman', 24), anchor=CENTER
            )
name_entry = Entry(
                base, textvariable=folder_name,
                font=('Times New Roman', 18))


#but_first.grid(row=0, column=0)
#lab_coun.grid(row=0, column=1)
but_first.pack(fill=BOTH, expand=TRUE)
lab_coun.pack(fill=BOTH, expand=TRUE)
name_entry.pack(fill=BOTH, expand=TRUE)

base.mainloop()
