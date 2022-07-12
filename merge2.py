#!python
from unicodedata import name
import openpyxl as xl
from tkinter import *
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfile

COPY_COLS = 3
status = 1

base = Tk()
base.title("Мержит файлы в один, за раз")
base.resizable(False, False)
base.geometry("600x450")


def save_file():
    global folder
    folder = asksaveasfile(filetypes=(("excel files","*.xlsx"), ("All files","*.*")), defaultextension=("excel files","*.xlsx"))

def browsefunc():    
    global status
    
    r = 0

    filename1 = askopenfilenames(filetypes=(("excel files","*.xlsx"), ("All files","*.*")))
                
    while r < len(filename1)-1:
        if status:
            wb1 = xl.load_workbook(filename1[r+1]) 
            ws1 = wb1.worksheets[0]
            rows_wb1 = ws1.max_row

            wb2 = xl.load_workbook(filename1[r]) 
            ws2 = wb2.worksheets[0]
            rows_wb2 = ws2.max_row      

            i = 1    
            for c in range(1, COPY_COLS+1):
                for i in range(0, rows_wb1):
                    ws2.cell(rows_wb2+i+1, c).value = ws1.cell(i+3, c).value
            r += 1
            status = 0

            wb1.close    
            wb2.save(folder.name)
            wb2.close  
        else:
            wb1 = xl.load_workbook(filename1[r+1]) 
            ws1 = wb1.worksheets[0]
            rows_wb1 = ws1.max_row

            wb2 = xl.load_workbook(folder.name) 
            ws2 = wb2.worksheets[0]
            rows_wb2 = ws2.max_row      

            i = 1    
            for c in range(1, COPY_COLS+1):
                for i in range(0, rows_wb1):
                    ws2.cell(rows_wb2+i+1, c).value = ws1.cell(i+3, c).value
            r += 1
            status = 0

            wb1.close    
            wb2.save(folder.name)
            wb2.close

    print("Одна судьба у наших двух сердец:")
    print("Замрет мое — и твоему конец!")    

def equalfunc():
    file_from_where = askopenfilename(filetypes=(("excel files","*.xlsx"), ("All files","*.*")))
    wb1 = xl.load_workbook(file_from_where) 
    ws1 = wb1.worksheets[0]    
    
    file_where = askopenfilename(filetypes=(("excel files","*.xlsx"), ("All files","*.*")))
    wb2 = xl.load_workbook(file_where) 
    ws2 = wb2.worksheets[0]
    rows_wb2 = ws2.max_row                          
    
    i = 1
    for c in range(1, COPY_COLS+1):
        for i in range(0, rows_wb2):
            ws2.cell(i+5, c+4).value = ws1.cell(i+3, c).value

    for i in range(0, rows_wb2-4):
        ws2.cell(i+5, c+6).value = "=B{}-F{}".format(i+5, i+5)
        ws2.cell(i+5, c+7).value = "=C{}-G{}".format(i+5, i+5)    

    wb1.close    
    wb2.save(folder.name)  
    wb2.close

    print("Кто предает себя же самого —")
    print("Не любит в этом мире никого!")

but_save = Button(
            base, 
            text="Создание файла 'который'!",
            bg='black', fg='white', width=20, height=5, font=('Times New Roman', 18),
            justify='left', command=save_file)
but_prima = Button(
            base, 
            text="Выбор файлов 'мельких',\nзамержиться в 'который'!",
            bg='white', fg='black', width=20, height=5, font=('Times New Roman', 18), 
            justify='left', command=browsefunc
            )
but_secunda = Button(
            base, 
            text="Первый файл 'который', второй 'большой'.\nСохраниться 'который'!",
            bg='black', fg='white', width=20, height=5, font=('Times New Roman', 18), 
            justify='left', command=equalfunc
            )

but_save.pack(fill=BOTH, expand=TRUE)
but_prima.pack(fill=BOTH, expand=TRUE)
but_secunda.pack(fill=BOTH, expand=TRUE)

base.mainloop()
