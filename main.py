#!python
import openpyxl as xl

# COPY_ROWS = 94
COPY_COLS = 3  
FOLDER = 'H:/temp/'

# source          
filename1 = 'merged.xlsx'
wb1 = xl.load_workbook(FOLDER + filename1) 
ws1 = wb1.worksheets[0]                         

#destination
filename2 = 'Перечень координат характерных точек границ территории.xlsx'
wb2 = xl.load_workbook(FOLDER + filename2) 
ws2 = wb2.worksheets[0]                          

r = 0
i = 1
count = 0
while r < ws1.max_row:
    r += 1
    for c in range(1, COPY_COLS+1):
        for i in range(0,COPY_ROWS):
           ws2.cell(i+5,c+4).value = ws1.cell(i+3,c).value
    count += 1
    r += i

wb1.close    
wb2.save(FOLDER + "algorithm1.xlsx")  
wb2.close

print("{} blocks of {} rows by {} columns copied".format(count,COPY_ROWS,COPY_COLS))